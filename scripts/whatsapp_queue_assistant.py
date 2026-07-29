#!/usr/bin/env python3
"""Cola local de revisión para WhatsApp Web.

Abre cada conversación con el texto precargado. El usuario conserva siempre
el control del botón Enviar. No automatiza clics ni simula comportamiento humano.
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import tkinter as tk
import urllib.parse
import webbrowser
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def normalize_phone(value: object, default_country: str = "34") -> str:
    raw = re.sub(r"\D", "", str(value or ""))
    if raw.startswith("00"):
        raw = raw[2:]
    if len(raw) == 9:
        raw = default_country + raw
    return raw


def parse_clock(value: str) -> dt.time:
    return dt.datetime.strptime(value.strip(), "%H:%M").time()


class WhatsAppQueueApp:
    def __init__(self, root: tk.Tk, initial_file: Path | None = None) -> None:
        self.root = root
        self.root.title("Mira · Asistente de WhatsApp")
        self.root.geometry("940x720")
        self.root.minsize(760, 620)
        self.workbook = None
        self.sheet = None
        self.file_path: Path | None = None
        self.output_path: Path | None = None
        self.rows: list[dict] = []
        self.position = 0
        self.paused = True
        self.pending_after: str | None = None

        self.file_label = tk.StringVar(value="Selecciona el Excel de Mira")
        self.business_var = tk.StringVar(value="—")
        self.phone_var = tk.StringVar(value="—")
        self.progress_var = tk.StringVar(value="0 / 0")
        self.status_var = tk.StringVar(value="En pausa")
        self.start_var = tk.StringVar(value="09:30")
        self.end_var = tk.StringVar(value="19:30")
        self.interval_var = tk.IntVar(value=45)
        self.only_published_var = tk.BooleanVar(value=True)

        self._build_ui()
        if initial_file and initial_file.exists():
            self.load_excel(initial_file)

    def _build_ui(self) -> None:
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Title.TLabel", font=("Segoe UI", 22, "bold"))
        style.configure("Heading.TLabel", font=("Segoe UI", 12, "bold"))
        style.configure("Primary.TButton", font=("Segoe UI", 11, "bold"), padding=10)
        style.configure("Secondary.TButton", padding=9)

        outer = ttk.Frame(self.root, padding=24)
        outer.pack(fill="both", expand=True)
        ttk.Label(outer, text="Asistente de WhatsApp Mira", style="Title.TLabel").pack(anchor="w")
        ttk.Label(
            outer,
            text="Abre cada conversación con el mensaje preparado. Revisa el texto y pulsa Enviar personalmente en WhatsApp Web.",
            wraplength=860,
        ).pack(anchor="w", pady=(5, 18))

        file_bar = ttk.Frame(outer)
        file_bar.pack(fill="x")
        ttk.Button(file_bar, text="Seleccionar Excel", command=self.choose_excel, style="Primary.TButton").pack(side="left")
        ttk.Label(file_bar, textvariable=self.file_label).pack(side="left", padx=14)

        schedule = ttk.LabelFrame(outer, text="Horario y ritmo de revisión", padding=14)
        schedule.pack(fill="x", pady=18)
        ttk.Label(schedule, text="Desde").grid(row=0, column=0, sticky="w")
        ttk.Entry(schedule, textvariable=self.start_var, width=9).grid(row=1, column=0, padx=(0, 14), sticky="w")
        ttk.Label(schedule, text="Hasta").grid(row=0, column=1, sticky="w")
        ttk.Entry(schedule, textvariable=self.end_var, width=9).grid(row=1, column=1, padx=(0, 14), sticky="w")
        ttk.Label(schedule, text="Espera tras confirmar (segundos)").grid(row=0, column=2, sticky="w")
        ttk.Spinbox(schedule, from_=10, to=600, textvariable=self.interval_var, width=10).grid(row=1, column=2, padx=(0, 14), sticky="w")
        ttk.Checkbutton(schedule, text="Solo informes publicados", variable=self.only_published_var).grid(row=1, column=3, sticky="w")

        card = ttk.LabelFrame(outer, text="Contacto actual", padding=18)
        card.pack(fill="both", expand=True)
        top = ttk.Frame(card)
        top.pack(fill="x")
        ttk.Label(top, textvariable=self.business_var, style="Heading.TLabel").pack(side="left")
        ttk.Label(top, textvariable=self.progress_var).pack(side="right")
        ttk.Label(card, textvariable=self.phone_var).pack(anchor="w", pady=(5, 12))
        self.message_box = tk.Text(card, height=13, wrap="word", font=("Segoe UI", 11), padx=12, pady=12)
        self.message_box.pack(fill="both", expand=True)

        actions = ttk.Frame(outer)
        actions.pack(fill="x", pady=(16, 8))
        ttk.Button(actions, text="Abrir en WhatsApp Web", command=self.open_current, style="Primary.TButton").pack(side="left")
        ttk.Button(actions, text="Marcar enviado y siguiente", command=self.mark_sent).pack(side="left", padx=8)
        ttk.Button(actions, text="Omitir", command=self.skip_current).pack(side="left")
        self.pause_button = ttk.Button(actions, text="Reanudar", command=self.toggle_pause)
        self.pause_button.pack(side="right")
        ttk.Label(outer, textvariable=self.status_var).pack(anchor="w")
        ttk.Label(
            outer,
            text="Importante: utiliza esta cola únicamente con contactos empresariales pertinentes y respetando oposición, consentimiento y políticas de WhatsApp. El programa nunca pulsa Enviar.",
            foreground="#786F82",
            wraplength=860,
        ).pack(anchor="w", pady=(14, 0))

    def choose_excel(self) -> None:
        selected = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if selected:
            self.load_excel(Path(selected))

    def load_excel(self, path: Path) -> None:
        try:
            workbook = load_workbook(path)
            sheet = next(
                candidate for candidate in workbook.worksheets
                if {"teléfono", "texto junto mensaje"}.issubset(
                    {normalize_header(cell.value) for cell in candidate[1]}
                )
            )
        except (OSError, StopIteration, ValueError) as error:
            messagebox.showerror("No se pudo abrir", f"El Excel no tiene las columnas necesarias.\n\n{error}")
            return

        headers = {normalize_header(cell.value): index + 1 for index, cell in enumerate(sheet[1])}
        status_column = headers.get("estado whatsapp")
        date_column = headers.get("fecha whatsapp")
        if status_column is None:
            status_column = sheet.max_column + 1
            sheet.cell(1, status_column, "Estado WhatsApp")
        if date_column is None:
            date_column = max(sheet.max_column + 1, status_column + 1)
            sheet.cell(1, date_column, "Fecha WhatsApp")

        rows = []
        for row in range(2, sheet.max_row + 1):
            phone = normalize_phone(sheet.cell(row, headers["teléfono"]).value)
            message = str(sheet.cell(row, headers["texto junto mensaje"]).value or "").strip()
            report_status = str(sheet.cell(row, headers.get("estado informe", 0)).value or "") if headers.get("estado informe") else ""
            whatsapp_status = str(sheet.cell(row, status_column).value or "")
            if not phone or not message or whatsapp_status.lower() == "enviado":
                continue
            if self.only_published_var.get() and "publicado" not in report_status.lower():
                continue
            rows.append({
                "row": row,
                "business": str(sheet.cell(row, headers.get("nombre", 1)).value or "Negocio"),
                "phone": phone,
                "message": message,
                "status_column": status_column,
                "date_column": date_column,
            })

        self.workbook = workbook
        self.sheet = sheet
        self.file_path = path
        self.output_path = path.with_name(f"{path.stem}_seguimiento_whatsapp.xlsx")
        self.rows = rows
        self.position = 0
        self.paused = True
        self.file_label.set(path.name)
        self.status_var.set(f"{len(rows)} contactos preparados · en pausa")
        self._show_current()

    def _show_current(self) -> None:
        if not self.rows or self.position >= len(self.rows):
            self.business_var.set("Cola terminada")
            self.phone_var.set("—")
            self.progress_var.set(f"{len(self.rows)} / {len(self.rows)}")
            self.message_box.delete("1.0", "end")
            return
        item = self.rows[self.position]
        self.business_var.set(item["business"])
        self.phone_var.set(f"+{item['phone']}")
        self.progress_var.set(f"{self.position + 1} / {len(self.rows)}")
        self.message_box.delete("1.0", "end")
        self.message_box.insert("1.0", item["message"])

    def _inside_schedule(self) -> bool:
        try:
            start, end = parse_clock(self.start_var.get()), parse_clock(self.end_var.get())
        except ValueError:
            messagebox.showerror("Horario incorrecto", "Utiliza el formato HH:MM, por ejemplo 09:30.")
            return False
        now = dt.datetime.now().time()
        return start <= now <= end if start <= end else now >= start or now <= end

    def open_current(self) -> None:
        if self.paused:
            messagebox.showinfo("Cola en pausa", "Pulsa Reanudar antes de abrir el siguiente contacto.")
            return
        if not self._inside_schedule():
            self.status_var.set("Fuera del horario configurado")
            return
        if self.position >= len(self.rows):
            return
        item = self.rows[self.position]
        message = self.message_box.get("1.0", "end").strip()
        url = f"https://web.whatsapp.com/send?phone={item['phone']}&text={urllib.parse.quote(message)}"
        webbrowser.open(url, new=0)
        self.status_var.set("Conversación abierta · revisa y pulsa Enviar en WhatsApp Web")

    def mark_sent(self) -> None:
        self._mark("Enviado")
        if not self.paused and self.position < len(self.rows):
            delay = max(10, int(self.interval_var.get())) * 1000
            self.status_var.set(f"Siguiente contacto preparado en {delay // 1000} segundos")
            self.pending_after = self.root.after(delay, self.open_current)

    def skip_current(self) -> None:
        self._mark("Omitido")

    def _mark(self, status: str) -> None:
        if self.position >= len(self.rows) or not self.sheet or not self.workbook or not self.output_path:
            return
        item = self.rows[self.position]
        self.sheet.cell(item["row"], item["status_column"], status)
        self.sheet.cell(item["row"], item["date_column"], dt.datetime.now())
        try:
            self.workbook.save(self.output_path)
        except OSError as error:
            messagebox.showerror("No se pudo guardar", str(error))
            return
        self.position += 1
        self._show_current()
        self.status_var.set(f"{status}. Seguimiento guardado en {self.output_path.name}")

    def toggle_pause(self) -> None:
        self.paused = not self.paused
        if self.paused and self.pending_after:
            self.root.after_cancel(self.pending_after)
            self.pending_after = None
        self.pause_button.configure(text="Reanudar" if self.paused else "Pausar")
        self.status_var.set("En pausa" if self.paused else "Cola activa · el botón Enviar sigue siendo manual")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("excel", nargs="?", type=Path)
    args = parser.parse_args()
    root = tk.Tk()
    WhatsAppQueueApp(root, args.excel)
    root.mainloop()


if __name__ == "__main__":
    main()
