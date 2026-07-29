#!/usr/bin/env python3
"""Genera y publica informes Mira, y devuelve el Excel con sus enlaces."""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


TERMINAL = {"finished", "completed", "ready", "needs_review", "partially_completed", "failed", "cancelled"}
PUBLIC_GEO_USER_AGENT = "MiraBusinessReports/1.0 (public business competitor research)"
COMPETITOR_CACHE: dict[str, dict] = {}


def slugify(value: str) -> str:
    plain = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode().lower()
    return re.sub(r"(^-|-$)", "", re.sub(r"[^a-z0-9]+", "-", plain))[:72] or "tienda"


def normalized_header(value: object) -> str:
    plain = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode().lower()
    return re.sub(r"\s+", " ", plain).strip()


def request_json(url: str, payload: dict | None = None) -> dict:
    data = json.dumps(payload).encode() if payload is not None else None
    request = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json", "Accept": "application/json"})
    with urllib.request.urlopen(request, timeout=120) as response:
        return json.load(response)


def request_public_json(url: str, payload: bytes | None = None) -> object:
    request = urllib.request.Request(
        url,
        data=payload,
        headers={
            "Accept": "application/json",
            "Content-Type": "application/x-www-form-urlencoded",
            "User-Agent": PUBLIC_GEO_USER_AGENT,
        },
    )
    with urllib.request.urlopen(request, timeout=45) as response:
        return json.load(response)


def location_query(address: str) -> str:
    """Obtiene una localidad aproximada sin conservar coordenadas personales."""
    postal_city = re.search(r"\b(\d{5})\s+([^,;|]+)", address)
    if postal_city:
        return f"{postal_city.group(1)} {postal_city.group(2).strip()}"
    parts = [part.strip() for part in re.split(r"[,;|]", address) if part.strip()]
    useful = [
        part for part in parts
        if not re.fullmatch(r"[\d\s\-]+", part)
        and not re.search(r"\b(calle|carrer|avenida|avinguda|plaza|paseo|passeig|nº|numero)\b", part, re.I)
    ]
    return ", ".join(useful[-2:]) if useful else address.strip()


def discover_public_competitors(source: dict) -> dict:
    """Busca comercios de moda públicos en OSM; nunca rastrea perfiles personales."""
    query = location_query(source.get("address", ""))
    if not query:
        return {
            "status": "location_missing",
            "location": "",
            "businesses": [],
            "evidence": "No había una localidad verificable en el Excel; no se han inventado competidores.",
            "policy": "Solo se consultan fichas comerciales públicas de OpenStreetMap.",
        }
    cache_key = slugify(query)
    if cache_key not in COMPETITOR_CACHE:
        try:
            geocode_url = (
                "https://nominatim.openstreetmap.org/search?"
                + urllib.parse.urlencode({"q": query, "format": "jsonv2", "limit": 1})
            )
            geocoded = request_public_json(geocode_url)
            if not isinstance(geocoded, list) or not geocoded:
                raise ValueError("Localidad no encontrada")
            lat, lon = float(geocoded[0]["lat"]), float(geocoded[0]["lon"])
            overpass_query = f"""
                [out:json][timeout:30];
                (
                  nwr(around:6000,{lat},{lon})["shop"~"^(clothes|fashion|boutique)$"];
                  nwr(around:6000,{lat},{lon})["craft"="tailor"];
                );
                out center tags;
            """
            encoded = urllib.parse.urlencode({"data": overpass_query}).encode()
            result = request_public_json("https://overpass-api.de/api/interpreter", encoded)
            elements = result.get("elements", []) if isinstance(result, dict) else []
            public_businesses = []
            seen: set[str] = set()
            for element in elements:
                tags = element.get("tags", {})
                name = str(tags.get("name") or "").strip()
                if not name or slugify(name) in seen:
                    continue
                seen.add(slugify(name))
                public_businesses.append({
                    "name": name,
                    "category": tags.get("shop") or tags.get("craft") or "moda",
                    "website": tags.get("website") or tags.get("contact:website") or "",
                    "source": "OpenStreetMap",
                })
            COMPETITOR_CACHE[cache_key] = {
                "status": "verified" if public_businesses else "not_found",
                "location": query,
                "businesses": public_businesses[:30],
            }
        except (OSError, ValueError, urllib.error.URLError):
            COMPETITOR_CACHE[cache_key] = {
                "status": "unavailable",
                "location": query,
                "businesses": [],
            }
    cached = COMPETITOR_CACHE[cache_key]
    own_name = slugify(source.get("name", ""))
    competitors = [
        business for business in cached["businesses"]
        if slugify(business["name"]) != own_name
    ][:4]
    return {
        "status": "verified" if competitors else cached["status"],
        "location": cached["location"],
        "businesses": competitors,
        "evidence": (
            f"Se localizaron {len(competitors)} comercios públicos comparables alrededor de {cached['location']}."
            if competitors
            else f"No se pudieron confirmar comercios comparables alrededor de {cached['location']}."
        ),
        "policy": "Solo se consultan fichas comerciales públicas de OpenStreetMap; no se investigan personas ni cuentas privadas.",
    }


def build_outreach_message(source: dict, report: dict, competitors: dict, url: str) -> str:
    findings = [*report.get("findability", []), *report.get("sales", [])]
    actions = [str(item.get("action", "")).strip() for item in findings if item.get("action")]
    priority = actions[0] if actions else "hacer más claro el camino desde la visita hasta la compra"
    names = [item["name"] for item in competitors.get("businesses", [])[:2]]
    comparison = (
        f" Durante el análisis hemos visto que comparte zona o compite digitalmente con negocios como {' y '.join(names)}."
        if names
        else ""
    )
    return (
        f"Hola, hemos analizado la web de {source['name']} y hemos preparado un informe privado con mejoras "
        f"concretas para vender más. La primera oportunidad detectada es {priority.rstrip('.').lower()}."
        f"{comparison} Incluye una propuesta visual y un plan de redes de 30 días: {url} "
        "Si te interesa, podemos explicártelo en una llamada breve y dejar las mejoras preparadas durante agosto."
    )


def download(url: str, target: Path) -> bool:
    try:
        target.parent.mkdir(parents=True, exist_ok=True)
        with urllib.request.urlopen(url, timeout=120) as response, target.open("wb") as output:
            shutil.copyfileobj(response, output)
        return True
    except urllib.error.HTTPError as error:
        if error.code == 404:
            return False
        raise


def read_businesses(workbook_path: Path) -> list[dict]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    businesses: list[dict] = []
    by_website: dict[str, dict] = {}
    used_slugs: dict[str, int] = {}
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        headers = [normalized_header(value) for value in next(rows, [])]
        positions = {name: index for index, name in enumerate(headers)}
        web_column = positions.get("pagina web")
        name_column = positions.get("nombre")
        if web_column is None or name_column is None:
            continue
        for row_number, row in enumerate(rows, start=2):
            name = str(row[name_column] or "").strip()
            website = str(row[web_column] or "").strip()
            if not name or not website.startswith(("http://", "https://")):
                continue
            parsed = urllib.parse.urlsplit(website)
            host = (parsed.hostname or "").lower().removeprefix("www.")
            normalized = f"{host}{parsed.path.rstrip('/')}".lower()
            row_ref = {"sheet": sheet.title, "row": row_number}
            if normalized in by_website:
                by_website[normalized]["_rows"].append(row_ref)
                continue
            base_slug = slugify(name)
            used_slugs[base_slug] = used_slugs.get(base_slug, 0) + 1
            final_slug = base_slug if used_slugs[base_slug] == 1 else f"{base_slug}-{used_slugs[base_slug]}"
            business = {
                "name": name,
                "website": website,
                "phone": str(row[positions["telefono"]] or "").strip() if "telefono" in positions else "",
                "address": str(row[positions["direccion"]] or "").strip() if "direccion" in positions else "",
                "slug": final_slug,
                "_rows": [row_ref],
            }
            by_website[normalized] = business
            businesses.append(business)
    workbook.close()
    return businesses


def public_business(source: dict) -> dict:
    return {key: value for key, value in source.items() if not key.startswith("_")}


def public_url(base_url: str, slug: str) -> str:
    return f"{base_url.rstrip('/')}/{slug}"


def update_excel(
    source_path: Path,
    output_path: Path,
    completed: list[dict],
    failed: list[dict],
    base_url: str,
    published: bool,
) -> None:
    """Copia el original y modifica únicamente las columnas de resultados."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(source_path)
    for sheet in workbook.worksheets:
        headers = [normalized_header(cell.value) for cell in sheet[1]]
        positions = {name: index + 1 for index, name in enumerate(headers)}
        if "nombre" not in positions or "pagina web" not in positions:
            continue
        if "web creada" not in positions:
            positions["web creada"] = sheet.max_column + 1
            sheet.cell(1, positions["web creada"], "Web creada")
        if "texto junto mensaje" not in positions:
            positions["texto junto mensaje"] = max(sheet.max_column + 1, positions["web creada"] + 1)
            sheet.cell(1, positions["texto junto mensaje"], "Texto junto mensaje")
        sheet.column_dimensions[sheet.cell(1, positions["web creada"]).column_letter].width = 38
        sheet.column_dimensions[sheet.cell(1, positions["texto junto mensaje"]).column_letter].width = 90

    for item in completed:
        url = public_url(base_url, item["slug"])
        if published:
            message = item.get("outreachMessage") or (
                f"Hola, hemos preparado para {item['name']} un análisis personalizado con mejoras "
                f"para vender más y una demostración del probador virtual: {url}"
            )
        else:
            message = "Informe generado localmente. Pendiente de publicación; no enviar todavía."
        for ref in item.get("_rows", []):
            sheet = workbook[ref["sheet"]]
            positions = {normalized_header(cell.value): index + 1 for index, cell in enumerate(sheet[1])}
            link_cell = sheet.cell(ref["row"], positions["web creada"])
            link_cell.value = url if published else ""
            if published:
                link_cell.hyperlink = url
                link_cell.font = Font(
                    name=link_cell.font.name,
                    size=link_cell.font.sz,
                    bold=link_cell.font.bold,
                    italic=link_cell.font.italic,
                    color="0563C1",
                    underline="single",
                )
            message_cell = sheet.cell(ref["row"], positions["texto junto mensaje"], message)
            message_cell.alignment = Alignment(
                horizontal=message_cell.alignment.horizontal,
                vertical="top",
                wrap_text=True,
            )
            sheet.row_dimensions[ref["row"]].height = max(sheet.row_dimensions[ref["row"]].height or 15, 42)

    for item in failed:
        reason = "; ".join(item.get("errors", []))[:400] or f"No creado: {item.get('status', 'error')}"
        for ref in item.get("_rows", []):
            sheet = workbook[ref["sheet"]]
            positions = {normalized_header(cell.value): index + 1 for index, cell in enumerate(sheet[1])}
            sheet.cell(ref["row"], positions["web creada"], "")
            sheet.cell(ref["row"], positions["texto junto mensaje"], f"No creado: {reason}")

    temporary = output_path.with_suffix(".tmp.xlsx")
    workbook.save(temporary)
    workbook.close()
    os.replace(temporary, output_path)


def wait_for_deployment(base_url: str, slug: str, timeout_seconds: int) -> bool:
    """Confirma que el JSON está realmente visible en el dominio público."""
    deadline = time.time() + timeout_seconds
    expected = f"{base_url.rstrip('/')}/campaign/{urllib.parse.quote(slug)}.json"
    while time.time() < deadline:
        try:
            request = urllib.request.Request(expected, headers={"Accept": "application/json", "Cache-Control": "no-cache"})
            with urllib.request.urlopen(request, timeout=30) as response:
                if response.status == 200 and response.headers.get_content_type() in {"application/json", "text/plain"}:
                    json.load(response)
                    return True
        except (OSError, ValueError, urllib.error.URLError):
            pass
        time.sleep(15)
    return False


def build_frontend_business(source: dict, report: dict, result_path: str | None, competitors: dict) -> dict:
    try_on = report.get("tryOn")
    pages_analyzed = int(report.get("pagesAnalyzed") or 0)
    product = try_on.get("productName") if try_on else "una prenda real del catálogo"
    analysis_try_on = None
    if try_on and result_path:
        analysis_try_on = {
            "productName": product,
            "productPrice": try_on.get("productPrice") or "Precio no detectado",
            "personImage": "/tryon-models/female-neutral.png" if try_on.get("modelGender") != "male" else "/tryon-models/male-neutral.webp",
            "productImage": try_on["productImage"],
            "resultImage": result_path,
        }
    return {
        "slug": source["slug"],
        "name": source["name"],
        "sector": "tiendas",
        "website": source["website"],
        "phone": source["phone"],
        "city": source["address"],
        "analysis": {
            "summary": (
                f"Hemos analizado {pages_analyzed} páginas reales de {source['name']}. "
                "La propuesta se basa en los problemas y productos encontrados durante ese rastreo."
                if pages_analyzed > 0
                else
                f"Hemos revisado 1 URL principal de {source['name']}. "
                "La web no permitió completar el rastreo de páginas en este primer intento; completaremos el análisis antes de implementar y no mostraremos una puntuación artificial."
            ),
            "pagesAnalyzed": pages_analyzed,
            "seoScore": report["visibilityScore"] if pages_analyzed > 0 else None,
            "croScore": report["salesReadinessScore"] if pages_analyzed > 0 else None,
            "seo": report.get("findability", []),
            "cro": report.get("sales", []),
            "social": report.get("social", {
                "profiles": [],
                "status": "not_analyzed",
                "evidence": "La búsqueda de perfiles sociales todavía no se ha ejecutado para este informe.",
                "policy": "Solo se utilizan enlaces públicos publicados por el propio negocio.",
            }),
            "competitors": competitors,
            "captures": report.get("captures", []),
            "tryOn": analysis_try_on,
            "marketing": {
                "customerName": "Lucía",
                "customerInitials": "L",
                "scenarios": {
                    "interest": {
                        "signal": f"La clienta ha consultado varias veces {product}.",
                        "decision": "Resolver la duda de talla mientras el interés sigue activo.",
                        "message": f"Hola, Lucía. {product} sigue en tu selección. ¿Quieres que te ayudemos con la talla?",
                    },
                    "cart": {
                        "signal": f"Ha añadido {product} al carrito y no ha terminado la compra.",
                        "decision": "Ofrecer ayuda útil sin recurrir automáticamente a un descuento.",
                        "message": f"Hola, Lucía. Hemos guardado {product}. Si dudas con la talla o la entrega, te ayudamos.",
                    },
                    "return": {
                        "signal": "Ya compró anteriormente y ha aceptado recibir recomendaciones.",
                        "decision": "Mostrar novedades reales relacionadas con sus intereses.",
                        "message": "Hola, Lucía. Hemos preparado una selección de novedades que encaja con tu estilo. ¿Quieres verla?",
                    },
                },
            },
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("excel", type=Path, nargs="?")
    parser.add_argument("--url")
    parser.add_argument("--name")
    parser.add_argument("--phone", default="")
    parser.add_argument("--api", default="http://127.0.0.1:3000")
    parser.add_argument("--output", type=Path, default=Path("packages/web/public/campaign"))
    parser.add_argument("--status-file", type=Path)
    parser.add_argument("--updated-excel", type=Path)
    parser.add_argument("--public-base-url", default=os.environ.get("MIRA_PUBLIC_BASE_URL", "https://www.miraia.space"))
    parser.add_argument("--deployment-timeout", type=int, default=600)
    parser.add_argument("--limit", type=int, default=1000)
    parser.add_argument("--poll-seconds", type=int, default=15)
    parser.add_argument("--publish", action="store_true")
    args = parser.parse_args()

    if args.url:
        businesses = [{
            "name": args.name or urllib.parse.urlsplit(args.url).hostname or "Tienda",
            "website": args.url,
            "phone": args.phone,
            "address": "",
            "slug": slugify(args.name or urllib.parse.urlsplit(args.url).hostname or "tienda"),
            "_rows": [],
        }]
    elif args.excel:
        businesses = read_businesses(args.excel)[: args.limit]
    else:
        raise SystemExit("Indica un Excel o utiliza --url para una prueba individual.")
    if not businesses:
        raise SystemExit("El Excel no contiene webs válidas.")

    jobs = request_json(f"{args.api}/reports/batch", {"urls": [item["website"] for item in businesses], "mode": "quick"})["jobs"]
    by_url = {item["website"]: item for item in businesses}
    pending = {job["jobId"]: {**by_url[job["url"]], **job} for job in jobs}
    manifest = {"startedAt": time.time(), "total": len(pending), "completed": [], "failed": [], "publicBaseUrl": args.public_base_url}
    args.output.mkdir(parents=True, exist_ok=True)

    def write_status() -> None:
        status_manifest = {
            **manifest,
            "completed": [public_business(item) for item in manifest["completed"]],
            "failed": [public_business(item) for item in manifest["failed"]],
        }
        content = json.dumps(status_manifest, ensure_ascii=False, indent=2)
        (args.output / "_campaign-status.json").write_text(content, encoding="utf-8")
        if args.status_file:
            args.status_file.parent.mkdir(parents=True, exist_ok=True)
            args.status_file.write_text(content, encoding="utf-8")

    while pending:
        statuses = request_json(f"{args.api}/reports/status", {"jobIds": list(pending)})["jobs"]
        for status in statuses:
            if status["status"] not in TERMINAL:
                continue
            source = pending.pop(status["jobId"])
            if status["status"] in {"failed", "cancelled"}:
                manifest["failed"].append({**source, "status": status["status"], "errors": status.get("errors", [])})
                continue
            report = request_json(f"{args.api}/reports/{status['jobId']}/mira")
            result_public_path = None
            if (report.get("tryOn") or {}).get("resultAvailable"):
                asset = args.output.parent / "campaign-assets" / source["slug"] / "tryon-result.jpg"
                if download(f"{args.api}/reports/{status['jobId']}/tryon-result", asset):
                    result_public_path = f"/campaign-assets/{source['slug']}/tryon-result.jpg"
            competitors = discover_public_competitors(source)
            frontend = build_frontend_business(source, report, result_public_path, competitors)
            (args.output / f"{source['slug']}.json").write_text(json.dumps(frontend, ensure_ascii=False, indent=2), encoding="utf-8")
            outreach_message = build_outreach_message(
                source,
                report,
                competitors,
                public_url(args.public_base_url, source["slug"]),
            )
            manifest["completed"].append({
                **source,
                "jobId": status["jobId"],
                "durationMs": status.get("durationMs"),
                "outreachMessage": outreach_message,
            })
        manifest["remaining"] = len(pending)
        manifest["elapsedSeconds"] = round(time.time() - manifest["startedAt"])
        write_status()
        print(f"Completados: {len(manifest['completed'])}/{manifest['total']} · fallidos: {len(manifest['failed'])} · pendientes: {len(pending)}")
        if pending:
            time.sleep(args.poll_seconds)

    manifest["finishedAt"] = time.time()
    manifest["status"] = "completed" if not manifest["failed"] else "completed_with_errors"
    published = False
    if args.publish:
        try:
            subprocess.run(["git", "add", "packages/web/public/campaign", "packages/web/public/campaign-assets"], check=True)
            changed = subprocess.run(["git", "diff", "--cached", "--quiet"], check=False).returncode != 0
            if changed:
                subprocess.run(["git", "commit", "-m", f"Publicar campaña Mira ({len(manifest['completed'])} informes)"], check=True)
                subprocess.run(["git", "push"], check=True)
            if manifest["completed"] and wait_for_deployment(
                args.public_base_url,
                manifest["completed"][0]["slug"],
                args.deployment_timeout,
            ):
                published = True
            else:
                raise RuntimeError("GitHub recibió los archivos, pero el dominio no confirmó el despliegue a tiempo.")
        except (OSError, RuntimeError, subprocess.CalledProcessError) as error:
            manifest["publishError"] = str(error)
    manifest["published"] = published

    if args.excel and args.updated_excel:
        update_excel(
            args.excel,
            args.updated_excel,
            manifest["completed"],
            manifest["failed"],
            args.public_base_url,
            published,
        )
        manifest["updatedExcel"] = str(args.updated_excel.resolve())
        manifest["updatedExcelReady"] = True
    write_status()


if __name__ == "__main__":
    main()
